"""PDF Table Extractor - Extract tables to Excel."""

import asyncio
import io
from pathlib import Path
from typing import List

from app.processors.base import BaseProcessor


class PDFTableExtractor(BaseProcessor):
    """Extract tables from PDF and save to Excel."""

    def _parse_pages(self, pages: str, total_pages: int) -> str:
        """Convert page specification to tabula format."""
        if pages.lower() == "all":
            return "all"
        return pages

    async def execute(
        self,
        input_path: Path,
        output_path: Path,
        pages: str = "all",
    ) -> Path:
        """
        Extract tables from PDF to Excel.

        Uses tabula-py for table extraction and openpyxl for Excel output.

        Args:
            input_path: Path to the input PDF
            output_path: Path for the Excel output (.xlsx)
            pages: Pages to extract tables from ("all" or "1,3,5-7")

        Returns:
            Path to the Excel file
        """
        try:
            import tabula
            import pandas as pd
        except ImportError:
            # Fallback to PyMuPDF if tabula not available
            return await self._extract_with_pymupdf(input_path, output_path, pages)

        # Run blocking tabula operations in thread pool
        def _extract_and_save():
            tables = tabula.read_pdf(
                str(input_path),
                pages=self._parse_pages(pages, 0),
                multiple_tables=True,
                pandas_options={'header': None},
            )
            return tables

        tables = await asyncio.to_thread(_extract_and_save)

        if not tables:
            # No tables found, try alternative extraction
            return await self._extract_with_pymupdf(input_path, output_path, pages)

        # Save to Excel in thread pool
        output_path = output_path.with_suffix(".xlsx")

        def _save_excel():
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for idx, table in enumerate(tables):
                    sheet_name = f"Table_{idx + 1}"
                    table.to_excel(writer, sheet_name=sheet_name, index=False)

        await asyncio.to_thread(_save_excel)

        return output_path

    async def _extract_with_pymupdf(
        self,
        input_path: Path,
        output_path: Path,
        pages: str,
    ) -> Path:
        """Fallback extraction using PyMuPDF."""

        def _extract():
            import fitz  # PyMuPDF
            import openpyxl

            doc = fitz.open(str(input_path))
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Extracted_Data"

            current_row = 1

            # Parse pages
            if pages.lower() == "all":
                page_range = range(len(doc))
            else:
                page_range = []
                for part in pages.split(","):
                    part = part.strip()
                    if "-" in part:
                        start, end = part.split("-")
                        page_range.extend(range(int(start) - 1, int(end)))
                    else:
                        page_range.append(int(part) - 1)

            for page_num in page_range:
                if page_num >= len(doc):
                    continue

                page = doc[page_num]
                tables = page.find_tables()

                if tables:
                    for table in tables:
                        ws.cell(row=current_row, column=1, value=f"Page {page_num + 1}")
                        current_row += 1

                        for row_data in table.extract():
                            for col_idx, cell_value in enumerate(row_data, start=1):
                                ws.cell(row=current_row, column=col_idx, value=cell_value or "")
                            current_row += 1

                        current_row += 1

            doc.close()

            out = output_path.with_suffix(".xlsx")
            wb.save(out)
            return out

        return await asyncio.to_thread(_extract)
